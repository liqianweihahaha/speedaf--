from time import time
from typing import Dict
import requests
import openpyxl
import json
import threading
import random
def request(url, body, headers):
    re = requests.post(url=url,data=body.encode('utf-8'),headers=headers)
    return re

def print_mustParm(re , key, expect_vaule, key1 = None, key2 = None):
    result = json.loads(re.text)    
    if 'code' in result.keys() and result['code'] == 1001 and re.status_code == 400 and 'tracking_id' in result.keys() and 'message' in result.keys() :
        pass
    else:
        print('测试失败 返回和预期结果不正确')
        print('字段名',key, key1, key2, 'http_code:',re.status_code,'code:',result['code'] )




def test_require_param(http_URL):
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet1']
    body_data = sheet.cell(row=2, column=2).value
    test_name = sheet.cell(row=2, column=1).value
    expect_vaule = sheet.cell(row=2, column=5).value
    require_param = sheet.cell(row=2, column=3).value.split(",")
    url = http_URL + sheet.cell(row=2, column=4).value
    body_data_d = json.loads(body_data)
    body_data_aa = json.loads(body_data)
    
    for key, vaule in body_data_d.items():
        if key in require_param:
            body_data_aa = json.loads(body_data)
            body_data_aa.pop(key)
            body = json.dumps(body_data_aa, ensure_ascii=False, indent=4)
            re = request(url,body,headers)
            print_mustParm(re,key,expect_vaule=expect_vaule)

        #json格式二级校验
        if isinstance(vaule,dict): 
            for key1,vaule1 in vaule.items():
                if key1 in require_param: 
                    body_data_aa = json.loads(body_data)
                    body_data_aa[key].pop(key1)
                    body = json.dumps(body_data_aa,ensure_ascii=False,indent=4)
                    re = request(url,body,headers)
                    print_mustParm(re,key,key1=key1,expect_vaule=expect_vaule)

                #json格式三级校验
                # if isinstance(vaule1,dict): 
                #     for key2 in vaule1.keys():
                #         if key2 in require_param: 
                #             body_data_aa = json.loads(body_data)
                #             body_data_aa[key][key1].pop(key2)
                #             body = json.dumps(body_data_aa,ensure_ascii=False,indent=4)
                #             re = request(url,body,headers)
                #             print_mustParm(re,key=key,key1=key1,key2=key2,expect_vaule=expect_vaule)
                            

   
def order_test(data):
    UAT_URL = 'http://47.241.40.42:8480/open-api/wish/order/asn'
    headers = {'content-type': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36'}
    body = json.dumps(data,ensure_ascii=False,indent=4)
    t = time() * 1000
    re = requests.post(url=UAT_URL,data=body.encode('utf-8'),headers=headers)
    t1 = time() * 1000 -  t
    result = json.loads(re.text)  
    if re.status_code == 200  and 'label_base64' in result.keys():
        # print(result['code'],result['message'], result['tracking_id'])
        print(result)
        pass
        
        
    else:
        # print('返回错误了')
        print('请求时间',t1,'报错的客户单号',data['trackingId'], result)


#Excel路径
file_path = r'E:\工作\wish对接\wish接口test.xlsx'
#请求的地址
online_URL = 'https://apis.test.speedaf.com/'
# UAT_URL = 'http://10.240.5.18:8480/'
headers = {'content-type': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36'}

if __name__ == "__main__":
    # test_require_param(UAT_URL)

    for i in range(1):
        data = {
            "otype":"8551-1",
            "sname":"GD_USPS_FC",
            "stype":1,
            "apiKey":"B2QoiMnuzjP7KrvI",
            "isTest":"true",
            "pickup_type":1,
            "parcel":{
                "weight":0.45,
                "priceUnit":"dollar",
                "categoryEn":"",
                "hasBattery":"false",
                "weightUnit":"kg",
                "productList":[
                    {
                        "sku":"561f48726475cc5873cda03e",
                        "value":1.2,
                        "weight":1,
                        "imageUrl":"https://www.wish.com/c/561f48726475cc5873cda03e",
                        "quantity":1,
                        "categoryEn":"",
                        "hasBattery":"false",
                        "descriptionEn":"Yoga pants",
                        "originCountry":"CN",
                        "descriptionLocal":"瑜伽裤",
                        "originCountryCode":"CHN"
                    }
                ],
                "declareValue":1.2,
                "descriptionEn":"Yoga pants",
                "dimensionUnit":"cm",
                "priceCurrency":"usd",
                "descriptionLocal":"瑜伽裤",
                "intrinsic_value":"11"
            },
            "pickUp":{
                "phone":"13763813216",
              
                "addressEn":{
                    "city":"Zhongshanshi",
                    "name":"Zhoushisen",
                    "country":"China",
                    "province":"Guangdongsheng",
                    "streetAddress1":"Guangdongshengzhongshanshisanjiaozhenjinsandadaodongsanshijiuhaozhiyililangdacangku（chengtaiduimian）"
                },
                "countryCode":"CHN",
                "addressLocal":{
                    "city":"中山市",
                    "name":"周世森",
                    "country":"中国",
                    "province":"广东省",
                    "streetAddress1":"广东省中山市三角镇金三大道东三十九号之一利朗达仓库（诚泰对面）"
                }
            },
            "sender":{
                "phone":"15989136708",
                "email":"",
                "zipCode":"310000",
                "addressEn":{
                    "city":"Shenzhen",
                    "name":"Baiying shenzhen trading company limited",
                    "country":"China",
                    "province":"Guangdong",
                    "streetAddress1":"Floor 5, building a, no.10 of east area weikangde industrial park shangxue technology city longgang"
                },
                "countryCode":"CHN",
                "addressLocal":{
                    "city":"Shenzhen",
                    "name":"BAIYING SHENZHEN TRADING COMPANY LIMITED",
                    "country":"中国",
                    "province":"Guangdong",
                    "streetAddress1":"Floor 5, Building A, No.10 of East Area Weikangde Industrial Park Shangxue Technology City Longgang"
                }
            },
            "isTaxed":"false",
            "receiver":{
                "phone":"198308888",
                "zipCode":"94016",
                "addressEn":{
                    "city":"San Francisco",
                    "name":"Carlota lopez",
                    "country":"Nigeria",
                    "province":"CA",
                    "streetAddress1":"54 ann st apt#1"
                },
                "countryCode":"NGA",
                "addressLocal":{
                    "city":"San Francisco",
                    "name":"Carlota Lopez",
                    "country":"Nigeria",
                    "province":"CA",
                    "streetAddress1":"54 ann st apt#1"
                }
            },
            "carryType":2,
            "orderTime":"2018-05-31T12:45:30.962Z",
            "timestamp":"2018-06-01T03:41:13.285Z",
            "pickupType":1,
            "trackingId":"WOSP024921274599NGA",
            "carrierCode":111,
            "requireLabel":"true",
            "wishpostSname":"GD_USPS_FC",
            "paymentAccount":{
                "id":"591185160e36e6d296fbddaf",
                "email":"shiqi@wish.com",
                "username":"shiqi",
                "contactName":"shiqi",
                "phoneNumber":"12345",
                "merchantPayVat":"false"
            },
            "wishpostServiceType":"3PL",
            "wishpostBusinessType":"GENERAL_LOGISTICS",
            "source_warehouse":"1",
            "destination_warehouse":"1",
            "user_custom_description":"sss",
            "paid_with_wish":"true",
            "return_info":{
                "return_action_in_country":0,
                "return_action_out_country":0,
                "return_address_in_country":{
                    "phone":"",
                    "mobile":"",
                    "country_code":"",
                    "address_en":{
                        "country":"",
                        "province":"",
                        "city":"",
                        "name":"",
                        "street_address1":""
                    },
                    "address_local":{
                        "country":"",
                        "province":"",
                        "city":"",
                        "name":"",
                        "street_address1":""
                    }
                },
                "return_address_out_country":{
                    "phone":"",
                    "mobile":"",
                    "country_code":"",
                    "address_en":{
                        "country":"",
                        "province":"",
                        "city":"",
                        "name":"",
                        "street_address1":"jkfsjljf"
                    },
                    "address_local":{
                        "country":"",
                        "province":"",
                        "city":"",
                        "name":"",
                        "street_address1":""
                    }
                }
            }
        } 

        customOrderNo = random.randint(12378743987,1237874398799)
        data["trackingId"] = str(customOrderNo)
        print(data["trackingId"])
        for k in range(1):
            t= threading.Thread(target=order_test,args=(data,))
            t.start()